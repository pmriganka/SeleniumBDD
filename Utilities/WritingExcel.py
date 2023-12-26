import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\pmrig\\OneDrive\\Desktop\\PDevelopment\\Development\\Selenium+BDD\\SeleniumBDD\\Utilities\\testdata.xlsx")
sheet = workbook["LoginTest"]
sheet.cell(row=5, column=1).value="mrigankap707@gmail.com"
sheet.cell(row=5, column=2).value="123456"
workbook.save("C:\\Users\\pmrig\\OneDrive\\Desktop\\PDevelopment\\Development\\Selenium+BDD\\SeleniumBDD\\Utilities\\testdata.xlsx")