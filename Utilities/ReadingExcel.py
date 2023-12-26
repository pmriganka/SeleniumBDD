import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\pmrig\\OneDrive\\Desktop\\PDevelopment\\Development\\Selenium+BDD\\SeleniumBDD\\Utilities\\testdata.xlsx")
sheet = workbook["LoginTest"]
totalrows = sheet.max_row
totalcols = sheet.max_column
print("Total rows are :", str(totalrows)," Total columns are :", str(totalcols))
print(sheet.cell(row=2, column=1).value)

for rows in range(1, totalrows+1):
    for cols in range(1, totalcols+1):
        print(sheet.cell(row=rows, column=cols).value, end=" ")
    print()