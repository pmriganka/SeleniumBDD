import openpyxl

def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column

def getCellData(path, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row = rowNum,column = colNum).value
    
def setCellData(path, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row = rowNum,column = colNum).value = data
    workbook.save(path)
    print("Data has been saved")

path="C:\\Users\\pmrig\\OneDrive\\Desktop\\PDevelopment\\Development\\Selenium+BDD\\SeleniumBDD\\Utilities\\testdata.xlsx"
sheetName = "LoginTest"
rows = getRowCount(path, sheetName)
print("Total rows : ",rows)
columns = getColCount(path, sheetName)
print("Total Columns : ",columns)
celldata = getCellData(path, sheetName, 4, 2)
print(celldata)
setCellData(path, sheetName, 4, 2, "sfasadfasdfsa")